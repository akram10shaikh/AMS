import csv

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, HttpResponseForbidden
from openpyxl import Workbook
from openpyxl.styles import Font
from .models import Player, Player_Group, CampTournament, CampActivity, Program, AssignedProgram, WorkoutData, Injury, \
    MedicalDocument
from player_app.forms import *
from django.http import JsonResponse
import json
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, get_user_model
from django.contrib.auth import logout
from django.contrib import messages
from openpyxl import load_workbook
from django.core.exceptions import ValidationError
from accounts.models import Organization, Staff
from form.forms import AssignForm
from form.models import FormAssignment
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist


# from ..app.models import DailyWellnessForm
# from player_app.ams.app.models import DailyWellnessForm


# View to list all players
def player_list(request):
    # Get all organizations for the superadmin
    organizations = Organization.objects.all()

    # Get the logged-in user's organization (if not a superuser)
    selected_organization = None
    players = []

    if request.user.is_superuser:
        # If superuser, allow organization selection
        selected_organization_id = request.GET.get('organization')
        if selected_organization_id:
            selected_organization = get_object_or_404(Organization, id=selected_organization_id)
            players = Player.objects.filter(organization=selected_organization)
    else:
        # Handle case where the staff user has no assigned organization
        try:
            if request.user.organization:  # Ensure organization exists
                selected_organization = request.user.organization
                players = Player.objects.filter(organization=selected_organization)
            else:
                players = []  # No players to show if no organization
        except ObjectDoesNotExist:
            players = []  # Handle missing organization safely

    # Pass organizations, selected organization, and players to the template
    return render(request, 'player_app/player_list.html', {
        'organizations': organizations,
        'selected_organization': selected_organization,
        'players': players,
    })


# View to display a single player's details
def player_detail(request, pk):
    player = get_object_or_404(Player, pk=pk)
    return render(request, 'player_app/player_detail.html', {'player': player})


from django.contrib.auth import get_user_model


def player_create(request):
    if request.method == 'POST':
        form = PlayerForm(request.POST, request.FILES)
        if form.is_valid():
            ins = form.save(commit=False)

            # Set the organization automatically if the user is an OrganizationAdmin
            if request.user.role == "OrganizationAdmin":
                # Automatically assign the logged-in organization's instance to this player
                organization = get_object_or_404(Organization, user=request.user)
                ins.organization = organization
            elif request.user.role == "Staff":
                # Fetch the organization of the logged-in Staff user
                staff = get_object_or_404(Staff, user=request.user)
                organization = staff.organization  # Assign the organization tied to the Staff user
                ins.organization = organization

            email = request.POST.get('email')
            firstname = request.POST.get('name').split(" ")[0]
            password = form.cleaned_data['password']

            User = get_user_model()
            user = User.objects.create(username=email, email=email, first_name=firstname)
            user.set_password(password)  # Set the password securely
            user.is_super_admin = False
            user.role = "Player"
            user.save()

            ins.user = user
            ins.save()
            messages.success(request, 'Player created successfully!')
            return redirect('player_list')
        else:
            messages.error(request, 'Please correct the errors below.')
            print("Form errors:", form.errors)
    else:
        form = PlayerForm()

    return render(request, 'player_app/player_form.html', {'form': form, 'title': 'Create Player'})


def player_update(request, pk):
    player = get_object_or_404(Player, pk=pk)
    original_organization = player.organization  # ✅ Store the original organization before update

    if request.method == 'POST':
        form = PlayerForm(request.POST, request.FILES, instance=player)
        files = request.FILES.getlist('documents')  # Get multiple uploaded documents

        if form.is_valid():
            password = form.cleaned_data.get('password')
            if password:
                user = player.user
                user.set_password(password)
                user.save()

            player = form.save(commit=False)  # ⛔️ Don't save yet

            # ✅ Preserve Organization: If missing, restore original organization
            if not player.organization:
                player.organization = original_organization

            player.save()  # ✅ Now save the updated player with the correct organization

            # Save uploaded medical documents
            for file in files:
                MedicalDocument.objects.create(player=player, document=file)

            messages.success(request, 'Player updated successfully!')
            return redirect('player_list')
        else:
            messages.error(request, 'Please correct the errors below.')
            print("Form errors:", form.errors)

    else:
        form = PlayerForm(instance=player)

    medical_documents = player.medical_documents.all()

    return render(request, 'player_app/player_form.html', {
        'form': form,
        'title': 'Update Player',
        'player': player,
        'medical_documents': medical_documents
    })


# View to delete an existing player
def player_delete(request, pk):
    player = get_object_or_404(Player, pk=pk)
    player.delete()
    messages.success(request, 'Player deleted successfully!')
    return redirect('player_list')




# View to list all players


def organization_player_list(request):
    if request.user.role == "Staff":
        org = request.user.staff.organization

    if request.user.role == "OrganizationAdmin":
        org = get_object_or_404(Organization, user=request.user)

    players = Player.objects.filter(organization=org)

    # Collect filter params
    age_categories = request.GET.getlist('age_category')
    handednesses = request.GET.getlist('handedness')
    roles = request.GET.getlist('role')
    player_statuses = request.GET.getlist('player_status')
    sort_gender = request.GET.get('sort') == 'gender'

    # Filtering logic
    if age_categories:
        players = players.filter(age_category__in=age_categories)
    if handednesses:
        players = players.filter(handedness__in=handednesses)
    if roles:
        players = players.filter(role__in=roles)
    if player_statuses:
        players = players.filter(player_status__in=player_statuses)

    # Calculate filters count including player_statuses
    filters_count = len(age_categories) + len(handednesses) + len(roles) + len(player_statuses)

    # Use your actual lowercase class attribute names for choices
    AGE_CHOICES = getattr(Player, "Age_category_choices", [])
    HAND_CHOICES = getattr(Player, "handedness_choices", [])
    ROLE_CHOICES = (
        Player.objects.filter(organization=org)
        .values_list('role', flat=True)
        .distinct()
        .order_by('role')
    )

    # Sorting by gender
    if sort_gender:
        gender_order = {'F': 0, 'M': 1, 'O': 2}
        players = sorted(
            players,
            key=lambda p: (
                gender_order.get(getattr(p, "gender", ""), 9),
                (getattr(p, "name", "") or '').lower()
            )
        )

    request_getlist = {k: request.GET.getlist(k) for k in request.GET}

    context = {
        'players': players,
        'AGE_CHOICES': AGE_CHOICES,
        'HAND_CHOICES': HAND_CHOICES,
        'ROLE_CHOICES': ROLE_CHOICES,
        'active_filters': {
            'age_category': age_categories,
            'handedness': handednesses,
            'role': roles,
            'player_status': player_statuses,
        },
        'filters_count': filters_count,
        'sort_gender': sort_gender,
        'request_getlist': request_getlist,
        'request_obj': request,
    }
    return render(request, "player_app/organization/organization_player_list.html", context)

@login_required
def organization_player_add(request):
    age_category_choices = getattr(Player, "Age_category_choices", [])
    return render(request, 'player_app/organization/organization_player_form.html', {'age_category_choices':age_category_choices})

@login_required
def player_create_view(request):
    if request.method == 'POST':
        # Extract fields manually from request.POST and request.FILES
        name = request.POST.get('name')
        image = request.FILES.get('image')  # File input
        email = request.POST.get('email')
        date_of_birth = request.POST.get('date_of_birth')
        primary_contact_number = request.POST.get('primary_contact_number')
        secondary_contact_number = request.POST.get('secondary_contact_number')
        gender = request.POST.get('gender')
        state = request.POST.get('state')
        district = request.POST.get('district')
        role = request.POST.get('role')
        batting_style = request.POST.get('batting_style')
        bowling_style = request.POST.get('bowling_style')
        handedness = request.POST.get('handedness')
        age_category = request.POST.get('age_category')
        guardian_name = request.POST.get('guardian_name')
        relation = request.POST.get('relation')
        guardian_mobile_number = request.POST.get('guardian_mobile_number')
        password = "admin"  # Or generate/set your own logic

        # Create Player/user instances and link
        User = get_user_model()

        # Basic Email uniqueness check to avoid duplicate users
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return render(request, 'player_app/organization/organization_player_form.html', {'title': 'Create Player'})

        user = User.objects.create(username=email, email=email, first_name=name.split(' ')[0])
        user.set_password(password)
        user.is_super_admin = False
        user.role = "Player"
        user.save()

        player = Player(
            name=name,
            email=email,
            date_of_birth=date_of_birth or None,
            primary_contact_number=primary_contact_number,
            secondary_contact_number=secondary_contact_number,
            gender=gender,
            state=state,
            district=district,
            role=role,
            batting_style=batting_style,
            bowling_style=bowling_style,
            handedness=handedness,
            age_category=age_category,
            guardian_name=guardian_name,
            relation=relation,
            guardian_mobile_number=guardian_mobile_number,
            user=user,
        )

        if image:
            player.image = image

        # Assign organization from user role as in your original logic
        if hasattr(request.user, 'role') and request.user.role == "OrganizationAdmin":
            organization = get_object_or_404(Organization, user=request.user)
            player.organization = organization
        elif hasattr(request.user, 'role') and request.user.role == "Staff":
            staff = get_object_or_404(Staff, user=request.user)
            player.organization = staff.organization

        player.save()

        # If you have M2M fields, handle here like player.m2mfield.set([...])

        PlayerActivityLog.objects.create(
            player=player,
            actor=request.user,
            action='created',
            details=f"Player '{player.name}' was created."
        )

        messages.success(request, 'Player created successfully!')
        return redirect('organization_player_list')

    
    return render(request, 'player_app/organization/organization_player_form.html', {'title': 'Create Player'})


from django.utils.timezone import now

@login_required
def organization_player_edit(request, pk):
    if request.user.role == "Staff":
        player = get_object_or_404(Player, pk=pk, organization=request.user.staff.organization)
        
        return redirect('organization_player_detail', pk=pk)
        
    org = get_object_or_404(Organization, user=request.user)
    player = get_object_or_404(Player, pk=pk, organization=org)
    user = player.user

    if request.method == 'POST':
        # Store old values for comparison
        old_data = {
            'name': player.name,
            'email': player.email,
            'primary_contact_number': player.primary_contact_number,
            'image':player.image.url if player.image else None,
            'date_of_birth': player.date_of_birth,
            'primary_contact_number': player.primary_contact_number,
            'secondary_contact_number': player.secondary_contact_number,
            'gender': player.gender,
            'state': player.state,
            'role': player.role,
            'batting_style': player.batting_style,
            'bowling_style': player.bowling_style,
            'handedness': player.handedness,
            'age_category': player.age_category,
            'guardian_name': player.guardian_name,
            'relation': player.relation,
            'guardian_mobile_number': player.guardian_mobile_number,
            
        }

        form = OrganizationPlayerFormUpdate(request.POST, request.FILES, instance=player)
        if form.is_valid():
            player = form.save()

            # Compare and prepare change log
            changes = []
            for field, old_value in old_data.items():
                new_value = getattr(player, field)
                if new_value != old_value:
                    changes.append(f"{field} changed from '{old_value}' to '{new_value}'")

            # Sync user first_name and email with updated player info
            user.first_name = player.name.split(' ')[0]
            user.email = player.email

            # Password update logic
            new_password = form.cleaned_data.get('new_password')
            if new_password:
                user.set_password(new_password)
                messages.info(request, "Password updated for this player.")
            user.save()

            # Log changes only if there are changes
            if changes:
                PlayerActivityLog.objects.create(
                    player=player,  # Use player, not injury
                    actor=request.user,
                    action='updated',
                    details="; ".join(changes)
                )

            messages.success(request, "Player updated successfully.")
            return redirect('organization_player_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = OrganizationPlayerFormUpdate(instance=player)

    return render(request, 'player_app/organization/organization_player_edit.html', {
        'form': form,
        'player': player,
        'title': 'Edit Player'
    })


@login_required
def organization_player_delete(request, pk):
    org = get_object_or_404(Organization, user=request.user)
    player = get_object_or_404(Player, pk=pk, organization=org)
    player.user.delete()  # OnDelete CASCADE deletes Player too
    messages.success(request, "Player deleted successfully!")
    return redirect('organization_player_list')

from django.db.models import Q

@login_required
def organization_player_detail(request, pk):
    player = get_object_or_404(Player, pk=pk)
    injuries = player.injuries.select_related('reported_by').all().order_by('-injury_date')  # You may want order
    documents = (
        player.medical_documents
        .filter(Q(view_option="profile") | Q(view_option="injury_profile"))
        .order_by('-date', '-uploaded_at')
    )
    # Dynamic status
    injury_status = "Injured" if player.injuries.filter(status='open').exists() else "Fit"
    participation_status = "Benched" if injury_status == "Injured" else "Available"

    # Doc upload logic
    if request.method == "POST":
        doc_form = MedicalDocumentForm(request.POST, request.FILES, player=player)
        if doc_form.is_valid():
            doc = doc_form.save(commit=False)
            doc.player = player
            doc.user = request.user                
            doc.save()

            # Log the upload
            MedicalActivityLog.objects.create(
                player=player,
                document=doc,
                user=request.user,
                activity_type='UPLOAD',
                description=f"{request.user.get_username()} uploaded medical document '{doc.title}'"
            )
            messages.success(request, "Medical document uploaded.")
            return redirect(request.path)
    else:
        doc_form = MedicalDocumentForm(player=player)

    context = {
        "player": player,
        "injuries": injuries,
        "documents": documents,
        "injury_status": injury_status,
        "participation_status": participation_status,
        "doc_form": doc_form,
    }
    return render(request, "player_app/organization/organization_player_detail.html", context)

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

@login_required
def organization_player_export(request):
    org = get_object_or_404(Organization, user=request.user)
    players = Player.objects.filter(organization=org)

    # Apply same filters as your list view!
    age_categories = request.GET.getlist('age_category')
    handednesses = request.GET.getlist('handedness')
    roles = request.GET.getlist('role')
    if age_categories:
        players = players.filter(age_category__in=age_categories)
    if handednesses:
        players = players.filter(handedness__in=handednesses)
    if roles:
        players = players.filter(role__in=roles)

    # If you want to keep column order/export identical, use the ALL_SWAP_COLS and FIXED_COLS in your JS.
    columns = [
        ("S.No", None),
        ("Player ID", "id"),
        ("Name", "name"),
        ("Email", "email"),
        ("D.O.B", "date_of_birth"),
        ("Primary Contact", "primary_contact_number"),
        ("Secondary Contact", "secondary_contact_number"),
        ("Gender", "get_gender_display"),
        ("State", "state"),
        ("Role", "role"),
        ("Batting Style", "batting_style"),
        ("Bowling Style", "bowling_style"),
        ("Handedness", "get_handedness_display"),
        ("Age Category", "get_age_category_display"),
        ("Guardian Name", "guardian_name"),
        ("Relation", "relation"),
        ("Guardian Mobile", "guardian_mobile_number")
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Players"

    # Header row
    ws.append([col[0] for col in columns])

    # Data rows
    for idx, player in enumerate(players, start=1):
        row = []
        for col_name, field in columns:
            if field is None:
                row.append(idx)  # S.No
            else:
                attr = getattr(player, field, "")
                if callable(attr):
                    val = attr()
                else:
                    val = attr
                row.append(val if val is not None else "")
        ws.append(row)

    # Fit width
    for i, col in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = min(length + 3, 35)

    # Return as download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=Players_Export.xlsx'
    wb.save(response)
    return response



#  Injury Management Views
from django.shortcuts import render, get_object_or_404, redirect
from .models import Organization, Player, Staff
from .forms import InjuryForm
from django.views.decorators.http import require_GET


from datetime import date, timedelta
from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.db.models import Count
from player_app.models import Injury, Organization, Player

from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.db.models import Count
from .models import Injury, Organization, Player

def organization_injury_list(request):
    # Identify organization
    if request.user.role == "Staff":
        org = request.user.staff.organization
    elif request.user.role == "OrganizationAdmin":
        org = get_object_or_404(Organization, user=request.user)
    else:
        org = None

    injuries = Injury.objects.filter(player__organization=org).select_related("player", "reported_by")
    today = date.today()

    # Filters from GET
    severity_vals = request.GET.getlist('severity')
    status_vals = request.GET.getlist('status')
    body_parts = request.GET.getlist('body_region')
    player_ids = request.GET.getlist('player_id')
    search_name = request.GET.get('name', '').strip()
    sort = request.GET.get('sort', '')

    month = request.GET.get('month')
    year = request.GET.get('year')
    range_filter = request.GET.get('range', '')
    season = request.GET.get('season')
    categories = request.GET.getlist('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    MONTH_CHOICES = [
        (1, "January"), (2, "February"), (3, "March"), (4, "April"),
        (5, "May"), (6, "June"), (7, "July"), (8, "August"),
        (9, "September"), (10, "October"), (11, "November"), (12, "December")
    ]

    # Generate seasons last 10 years (e.g. 2025-2026)
    current_season_start = today.year if today.month >= 4 else today.year - 1
    season_choices = [f"{y}-{y+1}" for y in range(current_season_start, current_season_start - 10, -1)]

    # Filtering logic
    if search_name:
        injuries = injuries.filter(player__name__icontains=search_name)
    else:
        if start_date and end_date:
            injuries = injuries.filter(injury_date__range=[start_date, end_date])
        elif range_filter == "last3":
            min_date = today - timedelta(days=90)
            injuries = injuries.filter(injury_date__gte=min_date)
        elif range_filter == "last6":
            min_date = today - timedelta(days=182)
            injuries = injuries.filter(injury_date__gte=min_date)
        elif month and year:
            injuries = injuries.filter(injury_date__month=int(month), injury_date__year=int(year))
        elif season and year:
            try:
                year_val = int(year)
                sy, ey = map(int, season.split('-'))
                start = date(sy, 4, 1)
                end = date(ey, 3, 31)
                injuries = injuries.filter(injury_date__gte=start, injury_date__lte=end, injury_date__year=year_val)
            except Exception:
                pass
        elif season:
            try:
                sy, ey = map(int, season.split('-'))
                start = date(sy, 4, 1)
                end = date(ey, 3, 31)
                injuries = injuries.filter(injury_date__gte=start, injury_date__lte=end)
            except Exception:
                pass

        if categories:
            injuries = injuries.filter(nature_of_injury__in=categories)
        if severity_vals:
            injuries = injuries.filter(severity__in=severity_vals)
        if status_vals:
            injuries = injuries.filter(status__in=status_vals)
        if body_parts:
            injuries = injuries.filter(affected_body_part__in=body_parts)
        if player_ids:
            injuries = injuries.filter(player__id__in=player_ids)

    # Sorting
    if sort == 'severity':
        injuries = injuries.order_by('severity')
    elif sort == 'date':
        injuries = injuries.order_by('injury_date')
    elif sort == 'status':
        injuries = injuries.order_by('status')

    # Stats counts
    injury_total = injuries.count()
    injury_open = injuries.filter(status='open').count()
    injury_closed = injuries.filter(status='closed').count()

    SEVERITY_CHOICES = Injury.SEVERITY_CHOICES
    STATUS_CHOICES = Injury.STATUS_CHOICES
    BODY_PART_CHOICES = list(
        Injury.objects.filter(player__organization=org)
        .values_list('affected_body_part', flat=True)
        .distinct()
        .exclude(affected_body_part__isnull=True)
        .exclude(affected_body_part__exact='')
    )
    BODY_PART_CHOICES = [(bp, bp.title()) for bp in BODY_PART_CHOICES if bp]

    PLAYER_CHOICES = Player.objects.filter(organization=org).values_list('id', 'name')

    CATEGORY_CHOICES = (
        Player.objects.filter(organization=org)
        .exclude(age_category__isnull=True)
        .exclude(age_category='')
        .values_list('age_category', flat=True)
        .distinct()
    )

    # Count active filters
    filters_count = (
        len(severity_vals) + len(status_vals) + len(body_parts) +
        len(player_ids) + len(categories)
    )
    if search_name:
        filters_count += 1
    if month or year:
        filters_count += 1
    if range_filter:
        filters_count += 1
    if season:
        filters_count += 1
    if start_date or end_date:
        filters_count += 1

    # Player participation stats
    # player_status_counts = injuries.values('player_status').annotate(count=Count('player_status'))
    # status_map = {"full participation": 0, "limited participation": 0, "no participation": 0}
    # for ps in player_status_counts:
    #     key = ps['player_status']
    #     if key in status_map:
    #         status_map[key] = ps['count']

    context = {
        'year_choices': list(range(2020, today.year + 1)),
        'injuries': injuries,
        'SEVERITY_CHOICES': SEVERITY_CHOICES,
        'STATUS_CHOICES': STATUS_CHOICES,
        'BODY_PART_CHOICES': BODY_PART_CHOICES,
        'PLAYER_CHOICES': PLAYER_CHOICES,
        'CATEGORY_CHOICES': CATEGORY_CHOICES,
        'season_choices': season_choices,
        'MONTH_CHOICES': MONTH_CHOICES,
        'injury_total': injury_total,
        'injury_open': injury_open,
        'injury_closed': injury_closed,
        'filters_count': filters_count,
        'today': today,
        'start_date': start_date,
        'end_date': end_date,
        'filter_month': month,
        'filter_year': year,
        'filter_range': range_filter,
        'filter_season': season,
        'filter_categories': categories,
       
        'active_filters': {
            'severity': severity_vals,
            'status': status_vals,
            'body_region': body_parts,
            'player_id': player_ids,
            'name': search_name,
            'category': categories,
            'month': month,
            'year': year,
            'range': range_filter,
            'season': season,
            'start_date': start_date,
            'end_date': end_date,
        },
        'request_getlist': {k: request.GET.getlist(k) for k in request.GET},
        'sort': sort,
    }

    return render(request, "player_app/organization/organization_injury_list.html", context)



@login_required
def activity_log_combined_view(request):
    # Fetch all medical logs
    medical_logs = MedicalActivityLog.objects.select_related('player', 'user', 'document').order_by('-timestamp')

    # Fetch all injury logs
    injury_logs = InjuryActivityLog.objects.select_related('injury', 'actor').order_by('-created_at')

    # Fetch all player logs
    player_logs = PlayerActivityLog.objects.select_related('player', 'actor').order_by('-created_at')

    logs = []

    for log in medical_logs:
        logs.append({
            'log_type': 'Medical',
            'time': log.timestamp,
            'user': log.user.username if log.user else '',
            'target': log.player.name,
            'action': log.activity_type,
            'desc': log.description,
        })

    for log in injury_logs:
        logs.append({
            'log_type': 'Injury',
            'time': log.created_at,
            'user': log.actor.username if log.actor else '',
            'target': str(log.injury),
            'action': log.action,
            'desc': log.details,
        })

    for log in player_logs:
        logs.append({
            'log_type': 'Player',
            'time': log.created_at,
            'user': log.actor.username if log.actor else '',
            'target': log.player.name,
            'action': log.action,
            'desc': log.details,
        })

    # Sort all logs by time descending
    logs.sort(key=lambda x: x['time'], reverse=True)

    return render(request, "player_app/organization/activity_log_combined.html", {
        'logs': logs,
    })

@login_required
def organization_create_injury(request):
    # Get the current user's organization
    if request.user.role == "Staff":
            organization = request.user.staff.organization

    if request.user.role == "OrganizationAdmin":
          organization = get_object_or_404(Organization, user=request.user)
   
    players_qs = Player.objects.filter(organization=organization)
    physios_qs = Staff.objects.filter(organization=organization, role__iexact='Physio')
    
    if request.method == 'POST':
        # Pass the filtered querysets to the form so the select fields are filtered
        form = InjuryForm(request.POST, players_qs=players_qs, physios_qs=physios_qs)
        if form.is_valid():
           
            injury = form.save()
            # Log injury creation
            InjuryActivityLog.objects.create(
                injury=injury,
                actor=request.user,
                action='created',
                details=f'Injury reported by {injury.reported_by} for player {injury.player}'
            )
            return redirect('organization_injury_list')  # Update to your desired redirect
        else:
            print(form.errors)# Print all form errors in console for debugging
    else:
        # GET: Create empty form with filtered choices
        form = InjuryForm(players_qs=players_qs, physios_qs=physios_qs)
    
    context = {
        'form': form,
    }
    return render(request, 'player_app/organization/injury_create.html', context)

@require_GET
def get_player_info(request, player_id):
    # AJAX view to return player info as JSON
    try:
        player = Player.objects.get(id=player_id)
        data = {
            'name': player.name,
            'gender': player.gender,
            'age': player.age,
            'email': player.email,
            'role': player.role,
            'bowling_style': player.bowling_style,
            'player_id': player.id,
            'date_of_birth': player.date_of_birth.strftime('%b %d, %Y') if player.date_of_birth else '',
            'contact_number': player.primary_contact_number,
            'handedness': player.handedness,
            'batting_style': player.batting_style,
            'district_cricket_association_id': player.state,
            'photo_url': player.image.url if player.image else 'https://randomuser.me/api/portraits/men/32.jpg',
        }
        return JsonResponse({'success': True, 'player': data})
    except Player.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Player not found.'}, status=404)


@login_required
def organization_injury_edit(request, pk):
    organization = get_object_or_404(Organization, user=request.user)
    injury = get_object_or_404(Injury, pk=pk, player__organization=organization)
    players_qs = Player.objects.filter(organization=organization)
    physios_qs = Staff.objects.filter(organization=organization, role__iexact='Physio')

    # Store original values for fields you want to track (add more fields as needed)
    original_data = {
        "name": injury.name,
        # "injury_type": injury.injury_type,
        "severity": injury.severity,
        "status": injury.status,
        "injury_date": injury.injury_date,
        "expected_date_of_return": injury.expected_date_of_return,
        "affected_body_part": injury.affected_body_part,
        "nature_of_injury": injury.nature_of_injury,
        # "body_segment": injury.body_segment,
        "player_status": injury.player_status,
        "venue": injury.venue,
        "notes": injury.notes,
        "reported_by_id": injury.reported_by_id,
        # Add other fields you want to track changes for
    }

    if request.method == "POST":
        form = InjuryFormUpdate(request.POST, instance=injury, players_qs=players_qs, physios_qs=physios_qs)
        if form.is_valid():
            updated_injury = form.save(commit=False)

            changed_fields = []
            for field, old_value in original_data.items():
                new_value = getattr(updated_injury, field)
                # For foreign keys, compare IDs
                if field.endswith('_id'):
                    if new_value != old_value:
                        changed_fields.append(f"{field[:-3]} changed from '{old_value}' to '{new_value}'")
                else:
                    if new_value != old_value:
                        changed_fields.append(f"{field} changed from '{old_value}' to '{new_value}'")

            updated_injury.save()

            # Only create log if something changed
            if changed_fields:
                InjuryActivityLog.objects.create(
                    injury=updated_injury,
                    actor=request.user,
                    action='updated',
                    details="; ".join(changed_fields)
                )
            else:
                pass

            return redirect('organization_injury_list')
    else:
        form = InjuryFormUpdate(instance=injury, players_qs=players_qs, physios_qs=physios_qs)
    return render(request, "player_app/organization/organization_injury_edit.html", {"form": form, "injury": injury})

from itertools import chain
from operator import attrgetter
@login_required
def organization_injury_detail(request, pk):
    organization = get_object_or_404(Organization, user=request.user)
    injury = get_object_or_404(Injury, pk=pk, player__organization=organization)
    player = get_object_or_404(Player, pk=injury.player.pk, organization=organization)

    # Fetch medical documents related to this injury
    medical_document = injury.documents.select_related('user').order_by('-date', '-uploaded_at')
    medical_documents = medical_document.filter(view_option__in=["injury_profile", "injury_only"])

    availability_form = PlayerAvailabilityForm(instance=injury)

    # Fetch activity logs: injury logs and medical logs combined and sorted by their date fields
    injury_logs = injury.activity_logs.select_related('actor').all()
    medical_logs = player.activity_logs.select_related('user', 'document').all()
    combined_logs = sorted(
        chain(injury_logs, medical_logs),
        key=lambda log: getattr(log, 'created_at', getattr(log, 'timestamp', None)),
        reverse=True
    )
    old_injury_status = injury.status
    old_player_statuss = injury.player.player_status
    oldabc = injury.player_status

    # print(f"Initial Injury Status: {old_injury_status}, Player Status: {old_player_status}")
    if request.method == 'POST':    
        if "availability_submit" in request.POST:
            availability_form = PlayerAvailabilityForm(request.POST, instance=injury)
            if availability_form.is_valid():
                old_status = injury.status
                old_player_status = injury.player_status
                availability_form.save()
                injury.refresh_from_db()  # <- refresh injury instance to get updated values
                new_status = injury.status
                new_player_status = injury.player.player_status
                
                status_changed = old_injury_status != new_status
                player_status_changed = old_player_statuss != new_player_status
                
                # Then build your message_parts and create log based on these refreshed values
                message_parts = []
                
                if status_changed:
                    message_parts.append(f"Status changed from '{old_injury_status}' to '{new_status}'")
                if player_status_changed:
                    message_parts.append(f"Participation changed from '{old_player_statuss}' to '{new_player_status}'")
                if message_parts:
                    InjuryActivityLog.objects.create(
                        injury=injury,
                        actor=request.user,
                        action="updated availability",
                        details='; '.join(message_parts)
                    )
                return redirect('organization_injury_detail', pk=injury.pk)

        else:
            form = MedicalDocumentFormN(request.POST, request.FILES, injury=injury)
            if form.is_valid():
                doc = form.save(commit=False)
                doc.injury = injury
                doc.player = player
                doc.user = request.user
                doc.save()
                MedicalActivityLog.objects.create(
                    player=player,
                    document=doc,
                    user=request.user,
                    activity_type='UPLOAD',
                    description=f"Uploaded document '{doc.title}'"
                )
                return redirect('organization_injury_detail', pk=injury.pk)
    else:
        form = MedicalDocumentFormN(injury=injury)

    return render(
        request,
        "player_app/organization/organization_injury_detail.html",
        {
            "injury": injury,
            "activity_logs": combined_logs,
            "player": player,
            "medical_documents": medical_documents,
            "form": form,
            "availability_form": availability_form,
        }
    )



@login_required
def organization_injury_delete(request, pk):
    organization = get_object_or_404(Organization, user=request.user)
    injury = get_object_or_404(Injury, pk=pk, player__organization=organization)
    if request.method == "POST":
        InjuryActivityLog.objects.create(
            injury=injury,
            actor=request.user,
            action='deleted',
            details='Injury record deleted'
        )
        injury.delete()
        return redirect("organization_injury_list")
    return render(request, "player_app/organization/organization_injury_confirm_delete.html", {"injury": injury})

@login_required
def organization_injury_export(request):
    organization = get_object_or_404(Organization, user=request.user)
    injuries = Injury.objects.filter(player__organization=organization).select_related("player", "reported_by")
    # _Apply filters as in your list view if you want filtered export_
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="injuries.csv"'
    writer = csv.writer(response)
    writer.writerow([
        "ID", "Player", "Severity", "Type", "Status", "Title", "Injury Date",
        "Expected Return", "Reported By", "Body Region", "Body Segment", "Venue", "Notes"
    ])
    for inj in injuries:
        writer.writerow([
            inj.id, inj.player.name, inj.get_severity_display(), inj.injury_type, inj.get_status_display(),
            inj.name, inj.injury_date, inj.expected_date_of_return, getattr(inj.reported_by, 'name', ''),
            inj.affected_body_part, inj.body_segment, inj.venue, inj.notes
        ])
    return response

# Camps & Tournaments views

def organization_camps_tournaments(request):
    """
    Displays a list of camps/tournaments.
    - Super Admins can view all camps.
    - Staff can see camps based on their permissions.
    - Users in an organization can see only camps from their organization.
    """

    # Super Admin: Sees all camps
    if request.user.is_superuser:
        organizations = Organization.objects.all()  # Allow filtering by organization
        selected_org = request.GET.get('organization')

        if selected_org:
            camps = CampTournament.objects.filter(organization_id=selected_org, is_deleted=False)
            male_players = Player.objects.filter(organization_id=selected_org, gender='Male')
            female_players = Player.objects.filter(organization_id=selected_org, gender='Female')
        else:
            camps = CampTournament.objects.filter(is_deleted=False)  # Show all camps
            male_players = Player.objects.filter(gender='Male')
            female_players = Player.objects.filter(gender='Female')

    # Staff with permission: Sees all camps
    elif hasattr(request.user, 'staff') and request.user.staff.view_camps_tournaments:
        org = request.user.staff.organization
        camps = CampTournament.objects.filter(is_deleted=False)
        male_players = Player.objects.filter(organization=org, gender='Male')
        female_players = Player.objects.filter(organization=org, gender='Female')

    # Regular users: See only camps in their organization
    elif hasattr(request.user, 'organization') and request.user.organization:
        org = request.user.staff.organization
        camps = CampTournament.objects.filter(organization=request.user.organization, is_deleted=False)
        male_players = Player.objects.filter(organization=org, gender='Male')
        female_players = Player.objects.filter(organization=org, gender='Female')

    # No Access: Deny access if the user doesn’t meet the criteria
    else:
        return HttpResponseForbidden(
            "You must belong to an organization or have the necessary permissions to view camps and tournaments.")
  


    return render(request, 'player_app/organization/organization_camps_tournaments.html', {
        'camps': camps,
        'male_players': male_players,
        'female_players': female_players,
        'organizations': organizations if request.user.is_superuser else None
    })

def organization_camp_detail(request, camp_id):
    """
    Displays details of a specific camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)
    return render(request, 'player_app/organization/organization_camp_detail.html', {'camp': camp})

def organization_edit_camp(request, camp_id):
    """
    Handles editing a specific camp/tournament, including participants.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)

    if request.method == 'POST':
        camp.name = request.POST.get('name')
        camp.camp_type = request.POST.get('camp_type')

        # Keep existing start date (only allow editing end date)
        end_date = request.POST.get('end_date')
        if end_date:
            camp.end_date = end_date

        camp.venue = request.POST.get('venue')

        # Update participants (only from the same organization)
        participant_ids = request.POST.getlist('participants')
        camp.participants.set(participant_ids)

        camp.save()

        # Log the update activity
        CampActivity.objects.create(
            camp=camp,
            action='updated',
            performed_by=request.user,
            details=f"Camp/Tournament '{camp.name}' was updated."
        )

        messages.success(request, "Camp/Tournament updated successfully!")
        return redirect('organization_camp_detail', camp_id=camp.id)

    # Get only players from the same organization
    participants = Player.objects.filter(organization=camp.organization)

    return render(request, 'player_app/organization/organization_edit_camp.html', {
        'camp': camp,
        'participants': participants
    })

@login_required
def organization_create_camp(request):
    """
    Create a new Camp/Tournament.
    """
    organizations = None
    players = Player.objects.none()  # Default to no players

    # Super Admin: Get all organizations
    if request.user.is_superuser:
        organizations = Organization.objects.all()
        players = Player.objects.all()  # Super Admins can see all players

    # Organization Admins: Check if user has an organization directly
    elif hasattr(request.user, "organization") and request.user.organization:
        organization = request.user.organization
        players = Player.objects.filter(organization=organization)

    # Staff Members: Ensure they have a staff profile before accessing
    elif hasattr(request.user, "staff") and request.user.staff:
        organization = request.user.staff.organization
        players = Player.objects.filter(organization=organization)

    else:
        return HttpResponseForbidden(
            "You must be a Super Admin, Organization Admin, or a Staff member to create a camp.")

    players_grouped = {'M': {}, 'F': {}}
    for player in players:
        gender = player.gender
        age_cat = player.age_category
        if gender in ['M', 'F']:
            if age_cat not in players_grouped[gender]:
                players_grouped[gender][age_cat] = []
            players_grouped[gender][age_cat].append({'id': player.id, 'name': player.name})

    players_grouped_json = json.dumps(players_grouped)

    if request.method == "POST":
        name = request.POST.get("name")
        camp_type = request.POST.get("camp_type")
        start_date = request.POST.get("start_date")
        gender = request.POST.get("gender")
        age_category = request.POST.get("age_category")
        venue = request.POST.get("venue")


        # Super Admin: Allow selecting an organization
        if request.user.is_superuser:
            organization_id = request.POST.get("organization")
            organization = get_object_or_404(Organization, id=organization_id)
        elif hasattr(request.user, "organization") and request.user.organization:
            # Organization Admins: Auto-set organization
            organization = request.user.organization
        elif hasattr(request.user, "staff") and request.user.staff:
            # Staff: Auto-set organization from staff profile
            organization = request.user.staff.organization
        else:
            return HttpResponseForbidden("You do not have permission to create a camp.")

        # Create the camp/tournament
        camp = CampTournament.objects.create(
            name=name,
            camp_type=camp_type,
            start_date=start_date,
            gender=gender,
            age_category=age_category,
            venue=venue,
            organization=organization,
            created_by=request.user
        )

        # Add participants (Only players from the same organization)
        selected_participants = request.POST.getlist("participants")
        valid_participants = players.filter(id__in=selected_participants)  # Ensure only valid participants
        camp.participants.set(valid_participants)

        messages.success(request, "Camp/Tournament created successfully!")
        return redirect("organization_camps_tournaments")  # Redirect after creation
    
    return render(request, "player_app/organization/organization_create_camp.html", {
        "organizations": organizations,  # Super Admin can select
        "players": players,    # Filtered players for the user
        "players_grouped_json": players_grouped_json,
    })

def organization_delete_camp(request, camp_id):
    """
    Allows authorized users to soft-delete a camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)  # Remove organization filtering
    camp.is_deleted = True
    camp.save()
    # Log the deletion activity
    CampActivity.objects.create(
        camp=camp,
        action='deleted',
        performed_by=request.user,
        details=f"Camp/Tournament '{camp.name}' was deleted."
    )
    messages.success(request, 'Camp/Tournament deleted successfully.')
    return redirect('organization_camps_tournaments')

from django.db.models import Avg
from django.db.models import Min, Max
from collections import defaultdict
# @login_required
# def test_dashboard(request):
#     # Get user's organization (adjust as per your user model)
#     user_organization = getattr(request.user, 'organization', None)
#     if not user_organization:
#         return render(request, 'player_app/organization/test_dashboard.html', {
#             'error_message': "Your account is not linked to any organization.",
#         })

#     # Get players in user's organization
#     players_in_org = Player.objects.filter(organization=user_organization)

#     # Handle new test result form (restrict player queryset to org players)
#     if request.method == 'POST':
#         add_form = TestAndResultForm(request.POST)
#         add_form.fields['player'].queryset = players_in_org
#         if add_form.is_valid():
#             add_form.save()
#             return redirect('test_dashboard')
#     else:
#         add_form = TestAndResultForm()
#         add_form.fields['player'].queryset = players_in_org

#     # Handle filter form (restrict player queryset to org players)
#     filter_form = TestSummaryFilterForm(request.GET or None)
#     filter_form.fields['player'].queryset = players_in_org

#     # Base queryset filtered by org players only
#     qs = TestAndResult.objects.select_related('player').filter(player__in=players_in_org).order_by('player__name', 'test', 'date', 'id')

#     # Apply filters if valid form submitted
#     if filter_form.is_valid():
#         if filter_form.cleaned_data.get('player'):
#             qs = qs.filter(player=filter_form.cleaned_data['player'])
#         if filter_form.cleaned_data.get('test'):
#             qs = qs.filter(test=filter_form.cleaned_data['test'])

#     # Group trials by (player_id, test)
#     player_test_trials = defaultdict(list)
#     for trial in qs:
#         key = (trial.player.id, trial.test)
#         player_test_trials[key].append(trial)

#     # Build summary rows without serial yet
#     summary_rows = []
#     for (player_id, test), trials in player_test_trials.items():
#         if not trials:
#             continue

#         # Last two trials chronologically
#         last_two_trials = trials[-2:] if len(trials) >= 2 else trials[-1:]

#         trial_1 = last_two_trials[0].trial if len(last_two_trials) == 2 else None
#         trial_2 = last_two_trials[-1].trial

#         best_trial = min(t.trial for t in trials)

#         last_trial_obj = last_two_trials[-1]
#         # Individual Average: mean of all trials for this player and test
        
#         indv_average = sum(t.trial for t in trials) / len(trials) if trials else None
#         group_average = TestAndResult.objects.filter(test=test).aggregate(Avg('trial'))['trial__avg']
        
#         summary_rows.append({
#             'player_name': last_trial_obj.player.name,
#             'test': test,
#             'last_date': last_trial_obj.date,
#             'last_phase': last_trial_obj.phase,
#             'trial_1': trial_1,
#             'trial_2': trial_2,
#             'best_trial': best_trial,
#             'indv_average': indv_average,
#             'group_average': group_average,
#         })

#     # Group rows by test
#     summary_by_test = defaultdict(list)
#     for row in summary_rows:
#         summary_by_test[row['test']].append(row)

#     # Assign serial numbers per test table starting at 1
#     for test_name, rows in summary_by_test.items():
#         for idx, row in enumerate(rows, start=1):
#             row['serial'] = idx

#     context = {
#         'add_form': add_form,
#         'form': filter_form,
#         'summary_by_test': dict(summary_by_test),
#     }
#     return render(request, 'player_app/organization/test_dashboard.html', context)

@login_required
def add_test_result(request):
    organization = getattr(request.user, 'organization', None)

    if request.method == 'POST':
        form = TestAndResultForm(request.POST, organization=organization)
        if form.is_valid():
            final_level_value = form.cleaned_data.get('best')
            result_id = NomativeData.objects.get(final_level=float(final_level_value))
            result = model_to_dict(result_id)
            instance = form.save(commit=False)
            instance.distance_covered = float(result["total_distance"])
            instance.predicted_vo2max = float(result["approximately_vo2max"])
            instance.save()
            return redirect('test_results_main') 
    else:
        form = TestAndResultForm(organization=organization)

    return render(request, 'player_app/organization/test_add.html', {'form': form})



from django.shortcuts import render
from django.db.models import Count, Q
from .models import Player, Injury, Staff
from django.contrib.auth.decorators import login_required
from itertools import chain
from collections import OrderedDict
from itertools import chain
from django.db.models import Count, Q

@login_required
def organization_dashboard_org(request):
    # Fetch filters from GET params or default to 'all'
    selected_category = request.GET.get('category', 'all')
    selected_gender = request.GET.get('gender', 'all')

    # Determine organization based on user role
    if request.user.role == "Staff":
        organization = request.user.staff.organization
    elif request.user.role == "OrganizationAdmin":
        organization = get_object_or_404(Organization, user=request.user)

        # Base queryset for players and injuries (unfiltered for cards)
    all_players = Player.objects.filter(organization=organization)
    all_injuries = Injury.objects.filter(player__organization=organization)

    # Cards definitions aligned with your Age_category_choices and Player model
    CATEGORY_CARDS = OrderedDict([
        ("B - U14",    {'gender': 'M', 'age_category': 'boys_under-15',    'label': "B - U14"}),
        ("B - U16",    {'gender': 'M', 'age_category': 'boys_under-16',    'label': "B - U16"}),  # Add to your model choices if missing
        ("B - U19",    {'gender': 'M', 'age_category': 'boys_under-19',    'label': "B - U19"}),
        ("B - U23",    {'gender': 'M', 'age_category': 'men_under-23',     'label': "B - U23"}),
        ("M - SENIOR", {'gender': 'M', 'age_category': 'men_senior',       'label': "M - SENIOR"}),
        ("W - U15",    {'gender': 'F', 'age_category': 'girls_under-15',   'label': "W - U15"}),
        ("G - U19",    {'gender': 'F', 'age_category': 'girls_under-19',   'label': "G - U19"}),
        ("W - U23",    {'gender': 'F', 'age_category': 'women_under-23',   'label': "W - U23"}),
        ("W - SENIOR", {'gender': 'F', 'age_category': 'women_senior',     'label': "W - SENIOR"}),
    ])

    category_cards = []
    for key, card_cat in CATEGORY_CARDS.items():
        players_qs = all_players.filter(
            gender=card_cat['gender'],
            age_category=card_cat['age_category']
        ).distinct()

        total_players = players_qs.values('id').distinct().count()
        full_participation = players_qs.filter(player_status__iexact="full participation").values('id').distinct().count()
        limited_participation = players_qs.filter(player_status__iexact="limited participation").values('id').distinct().count()
        no_participation = players_qs.filter(player_status__iexact="no participation").values('id').distinct().count()
        active_injury = all_injuries.filter(player__in=players_qs, status='open').values('player_id').distinct().count()

        
        category_cards.append({
            'label': card_cat['label'],
            'total': total_players,
            'full': full_participation,
            'limited': limited_participation,
            'none': no_participation,
            'active_injury': active_injury,
        })
    
    test_org = Player.objects.filter(organization=organization)
    
    test_filter = Player.objects.filter(player_status__iexact="full participation",age_category="men_senior" ,organization=organization).count()
    


    # --- Dashboard filter (all other data below can use the filtered players/injuries) ---
    players = all_players
    injuries = all_injuries.select_related('player', 'reported_by')

    # Filter by category if selected
    if selected_category != 'all':
        players = players.filter(age_category=selected_category)
        injuries = injuries.filter(player__age_category=selected_category)

    # Filter by gender if selected
    if selected_gender != 'all':
        players = players.filter(gender=selected_gender)
        injuries = injuries.filter(player__gender=selected_gender)

    # Total injuries and active injuries (filtered)
    total_injuries_count = injuries.count()
    active_injuries = injuries.filter(status='open')
    active_injuries_count = active_injuries.count()

    # Participation types (filtered)
    participation_counts = CampTournament.objects.filter(
        participants__in=players
    ).annotate(
        player_count=Count('participants', filter=Q(participants__in=players), distinct=True)
    ).order_by('-player_count')

    # Prefetch injuries (with notes)
    players = players.prefetch_related('injuries__reported_by')

    # Activity log setup
    medical_logs = MedicalActivityLog.objects.filter(player__in=players).select_related('player', 'user', 'document')
    injury_logs = InjuryActivityLog.objects.filter(injury__in=injuries).select_related('injury', 'actor')
    player_logs = PlayerActivityLog.objects.filter(player__in=players).select_related('player', 'actor')
    for log in medical_logs:
        log.log_type = 'medical'
    for log in injury_logs:
        log.log_type = 'injury'
    for log in player_logs:
        log.log_type = 'player'
    
    combined_logs = sorted(
        chain(medical_logs, injury_logs, player_logs),
        key=lambda log: getattr(log, 'timestamp', getattr(log, 'created_at', None)),
        reverse=True
    )

    

    context = {
        'selected_category': selected_category,
        'selected_gender': selected_gender,
        'players': players,
        'total_injuries_count': total_injuries_count,
        'active_injuries_count': active_injuries_count,
        'active_injuries': active_injuries,
        'participation_counts': participation_counts,
        'activity_logs': combined_logs,
        'category_cards': category_cards,  # <-- ADD THIS LINE
    }
    return render(request, 'player_app/organization/organization_dashboard.html', context)

def logout_user(request):
    """
    Logs out the user and redirects to the login page.
    """
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('home')  # Adjust to your login URL name
# --------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------


# View to export players to Excel

def export_players_to_excel(request):
    players = Player.objects.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Players'

    headers = [
        'name', 'aadhar number', 'batting style', 'bowling style', 'date of birth', 'email', 'gender',
        'guardian mobile number', 'guardian name', 'handedness', 'id card number', 'profile image',
        'medical certificates', 'primary contact number', 'relation', 'user_role', 'secondary contact number',
        'sports role', 'state', 'address', 'district', 'pincode', 'aadhar card upload', 'marksheets upload',
        'pan card upload', 'additional information', 'age category', 'allergies', 'disease', 'height',
        'nationality', 'position', 'team', 'weight'
    ]
    sheet.append(headers)

    # Make header row bold
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for player in players:
        sheet.append([
            player.name,
            player.aadhar_number,
            player.batting_style,
            player.bowling_style,
            player.date_of_birth,
            player.email,
            player.gender,
            player.guardian_mobile_number,
            player.guardian_name,
            player.handedness,
            player.id_card_number,
            player.image.url if player.image else 'N/A',
            player.medical_certificates.url if player.medical_certificates else 'N/A',
            player.primary_contact_number,
            player.relation,
            player.user_role,
            player.secondary_contact_number,
            player.sports_role,
            player.state,
            player.address,
            player.district,
            player.pincode,
            player.aadhar_card_upload.url if player.aadhar_card_upload else 'N/A',
            player.marksheets_upload.url if player.marksheets_upload else 'N/A',
            player.pan_card_upload.url if player.pan_card_upload else 'N/A',
            player.additional_information,
            player.age_category,
            player.allergies,
            player.disease,
            player.height,
            player.nationality,
            player.position,
            player.team,
            player.weight
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=players.xlsx'
    workbook.save(response)

    return response


def upload_medical_documents(request, player_id):
    player = get_object_or_404(Player, id=player_id)

    if request.method == 'POST':
        form = MultipleMedicalDocumentsForm(request.POST, request.FILES)

        if form.is_valid():
            files = request.FILES.getlist('documents')  # Get multiple files as a list

            for file in files:
                MedicalDocument.objects.create(player=player, document=file)  # Save each document

            messages.success(request, "Medical documents uploaded successfully!")
            return redirect('player_detail', pk=player.id)
        else:
            messages.error(request, "Error uploading documents.")

    form = MultipleMedicalDocumentsForm()
    return render(request, 'player_app/upload_medical_documents.html', {'form': form, 'player': player})


# View to import players to Excel

import logging

logger = logging.getLogger(__name__)

from django.db.models import Max
from accounts.models import CustomUser

from django.db import IntegrityError
from player_app.models import Player
from accounts.models import CustomUser
import pandas as pd
import logging

# Set up logging to track the process
logger = logging.getLogger(__name__)


def upload_file(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            df = pd.read_excel(excel_file)

            # Log column names and sample data for debugging
            logger.info(f"DataFrame columns: {df.columns}")
            logger.info(f"DataFrame head: \n{df.head()}")

            # Get the highest existing user_id (Player's ID) from the database
            max_user_id = Player.objects.aggregate(Max('user_id'))['user_id__max'] or 0
            next_user_id = max_user_id + 1  # This will be the next available user_id

            for _, row in df.iterrows():
                try:
                    user_id = row.get("user_id", None)

                    if not user_id:  # If user_id is missing, auto-increment from next available user_id
                        user_id = next_user_id
                        next_user_id += 1

                    # Check if the user with this user_id already exists in the CustomUser model
                    user_exists = CustomUser.objects.filter(id=user_id).exists()

                    while user_exists:  # If user_id already exists, try the next one
                        user_id = next_user_id
                        next_user_id += 1
                        user_exists = CustomUser.objects.filter(id=user_id).exists()

                    # If user doesn't exist, create a new user
                    user = CustomUser.objects.create(
                        id=user_id,  # Explicitly setting the user_id to avoid auto-increment conflict
                        username=row.get("name", ""),
                        password=row.get("password", ""),
                    )

                    # Prepare the data for creating Player instance
                    player_data = {
                        "user": user,
                        "name": row.get("name", ""),
                        "email": row.get("email", ""),
                        "primary_contact_number": row.get("primary_contact_number", ""),
                        "secondary_contact_number": row.get("secondary_contact_number", ""),
                        "date_of_birth": row.get("date_of_birth", None),
                        "pincode": row.get("pincode", ""),
                        "address": row.get("address", ""),
                        "nationality": row.get("nationality", ""),
                        "gender": row.get("gender", ""),
                        "state": row.get("state", ""),
                        "district": row.get("district", ""),
                        "role": row.get("role", ""),
                        "batting_style": row.get("batting_style", ""),
                        "bowling_style": row.get("bowling_style", ""),
                        "handedness": row.get("handedness", ""),
                        "aadhar_number": row.get("aadhar_number", ""),
                        "sports_role": row.get("sports_role", ""),
                        "id_card_number": row.get("id_card_number", ""),
                        "weight": float(row["weight"]) if pd.notna(row.get("weight")) else None,
                        "height": float(row["height"]) if pd.notna(row.get("height")) else None,
                        "age_category": row.get("age_category", ""),
                        "team": row.get("team", ""),
                        "position": row.get("position", ""),
                        "guardian_name": row.get("guardian_name", ""),
                        "relation": row.get("relation", ""),
                        "guardian_mobile_number": row.get("guardian_mobile_number", ""),
                        "disease": row.get("disease", ""),
                        "allergies": row.get("allergies", ""),
                        "additional_information": row.get("additional_information", ""),
                        # File handling for specific fields
                        "image": row["image"] if "image" in row and pd.notna(row["image"]) else None,
                        "medical_certificates": row[
                            "medical_certificates"] if "medical_certificates" in row and pd.notna(
                            row["medical_certificates"]) else None,
                        "aadhar_card_upload": row["aadhar_card_upload"] if "aadhar_card_upload" in row and pd.notna(
                            row["aadhar_card_upload"]) else None,
                        "pan_card_upload": row["pan_card_upload"] if "pan_card_upload" in row and pd.notna(
                            row["pan_card_upload"]) else None,
                        "marksheets_upload": row["marksheets_upload"] if "marksheets_upload" in row and pd.notna(
                            row["marksheets_upload"]) else None,
                    }

                    # Create the Player instance
                    Player.objects.create(**player_data)

                except IntegrityError as e:
                    logger.error(f"IntegrityError: {e} for row: {row}")
                    # Handle row-related issues here if necessary

            return redirect("player_list")
        else:
            logger.error("Form is invalid")
    else:
        form = UploadFileForm()

    return render(request, "player_list.html", {"form": form})


def download_blank_excel(request):
    # Get all fields from the Player model
    fields = Player._meta.get_fields()

    # List to store headers (field names)
    headers = []

    for field in fields:
        # Skip related model fields (e.g., ForeignKey, OneToMany)
        if field.is_relation:
            continue
        headers.append(field.name)

    # Create a new Workbook and select the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Players'

    # Append headers as the first row in the sheet
    sheet.append(headers)

    # Make header row bold
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Set the HTTP response for downloading the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=blank_players_template.xlsx'
    workbook.save(response)

    return response


# --------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------


def get_all_group_players(request):
    group_id = request.GET.get('group_id')
    group = get_object_or_404(Player_Group, pk=group_id)
    players = group.player_set.all().values('pk', 'name', 'image')
    return JsonResponse({'players': list(players)})


def get_all_players(request):
    players = Player.objects.all().values('pk', 'name', 'image')
    return JsonResponse({'players': list(players)})


# View to manage player groups (create/delete groups)
def manage_groups(request):
    groups = Player_Group.objects.all()
    players = Player.objects.all()
    group_form = GroupForm()

    if request.method == "POST":
        if 'create_group' in request.POST:
            group_form = GroupForm(request.POST)
            if group_form.is_valid():
                group = group_form.save()
                player_ids = request.POST.getlist('group_players')
                for player_id in player_ids:
                    player = Player.objects.get(pk=player_id)
                    player.players_in_groups.add(group)
                return redirect('manage_groups')  # Ensure redirect on success
        elif 'remove_player_from_group' in request.POST:
            group_id = request.POST.get('group_id')
            player_id = request.POST.get('player_id')
            group = get_object_or_404(Player_Group, pk=group_id)
            player = get_object_or_404(Player, pk=player_id)
            player.players_in_groups.remove(group)
            return redirect('manage_groups')  # Ensure redirect on success

    # Always return a response, whether GET or POST
    context = {
        'groups': groups,
        'players': players,
        'group_form': group_form
    }
    return render(request, 'player_app/player_group_manage.html', context)  # Ensure response is returned


# Adjust this to our main player_group_manage.html as a modal
# View to manage all groups (add/remove players)
def manage_all_groups(request):
    if request.method == 'POST':
        group_id = request.POST.get('group_id')
        action = request.POST.get('action')
        player_ids = request.POST.getlist('player_ids')

        group = get_object_or_404(Player_Group, pk=group_id)
        players = Player.objects.filter(pk__in=player_ids)

        if action == 'add':
            for player in players:
                player.players_in_groups.add(group)  # Add the player to the group
            messages.success(request, 'Players added to group successfully!')
        elif action == 'remove':
            for player in players:
                player.players_in_groups.remove(group)  # Remove the player from the group
            messages.success(request, 'Players removed from group successfully!')

        return redirect('manage_groups')  # Adjust this to your actual URL name for managing groups

    groups = Player_Group.objects.all()
    players = Player.objects.all()

    context = {
        'groups': groups,
        'players': players
    }

    return render(request, 'player_app/add_player_to_group.html', context)


def rename_group(request, group_id):
    group = get_object_or_404(Player_Group, id=group_id)

    if request.method == 'POST':
        new_name = request.POST.get('name')
        if new_name:
            group.name = new_name  # Update the name
            group.save()  # Save the changes
            return redirect('group_list')  # Redirect to the list of groups or another page

    return render(request, 'rename_group.html', {'group': group})


def delete_group(request, group_id):
    print(f"Delete group function called for group id: {group_id}")  # Debug statement
    group = get_object_or_404(Player_Group, pk=group_id)
    group.delete()
    return redirect('manage_groups')


def player_home(request, pk):
    player = get_object_or_404(Player, pk=pk)
    form_assignments = FormAssignment.objects.filter(player=player)
    forms = [assignment.form for assignment in form_assignments]
    return render(request, 'player_app/player_home.html', {'player': player, 'forms': forms})


from django.shortcuts import render
from django.http import HttpResponseForbidden
from .models import CampTournament, Organization


def camps_tournaments(request):
    """
    Displays a list of camps/tournaments.
    - Super Admins can view all camps.
    - Staff can see camps based on their permissions.
    - Users in an organization can see only camps from their organization.
    """

    # Super Admin: Sees all camps
    if request.user.is_superuser:
        organizations = Organization.objects.all()  # Allow filtering by organization
        selected_org = request.GET.get('organization')

        if selected_org:
            camps = CampTournament.objects.filter(organization_id=selected_org, is_deleted=False)
        else:
            camps = CampTournament.objects.filter(is_deleted=False)  # Show all camps

    # Staff with permission: Sees all camps
    elif hasattr(request.user, 'staff') and request.user.staff.view_camps_tournaments:
        camps = CampTournament.objects.filter(is_deleted=False)

    # Regular users: See only camps in their organization
    elif hasattr(request.user, 'organization') and request.user.organization:
        camps = CampTournament.objects.filter(organization=request.user.organization, is_deleted=False)

    # No Access: Deny access if the user doesn’t meet the criteria
    else:
        return HttpResponseForbidden(
            "You must belong to an organization or have the necessary permissions to view camps and tournaments.")

    return render(request, 'player_app/camps_tournaments.html', {
        'camps': camps,
        'organizations': organizations if request.user.is_superuser else None
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from player_app.models import CampTournament, Player
from accounts.models import Organization  # Ensure Organization is imported


@login_required
def create_camp(request):
    """
    Create a new Camp/Tournament.
    """
    organizations = None
    players = Player.objects.none()  # Default to no players

    # Super Admin: Get all organizations
    if request.user.is_superuser:
        organizations = Organization.objects.all()
        players = Player.objects.all()  # Super Admins can see all players

    # Organization Admins: Check if user has an organization directly
    elif hasattr(request.user, "organization") and request.user.organization:
        organization = request.user.organization
        players = Player.objects.filter(organization=organization)

    # Staff Members: Ensure they have a staff profile before accessing
    elif hasattr(request.user, "staff") and request.user.staff:
        organization = request.user.staff.organization
        players = Player.objects.filter(organization=organization)

    else:
        return HttpResponseForbidden(
            "You must be a Super Admin, Organization Admin, or a Staff member to create a camp.")

    if request.method == "POST":
        name = request.POST.get("name")
        camp_type = request.POST.get("camp_type")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        venue = request.POST.get("venue")

        # Super Admin: Allow selecting an organization
        if request.user.is_superuser:
            organization_id = request.POST.get("organization")
            organization = get_object_or_404(Organization, id=organization_id)
        elif hasattr(request.user, "organization") and request.user.organization:
            # Organization Admins: Auto-set organization
            organization = request.user.organization
        elif hasattr(request.user, "staff") and request.user.staff:
            # Staff: Auto-set organization from staff profile
            organization = request.user.staff.organization
        else:
            return HttpResponseForbidden("You do not have permission to create a camp.")

        # Create the camp/tournament
        camp = CampTournament.objects.create(
            name=name,
            camp_type=camp_type,
            start_date=start_date,
            end_date=end_date,
            venue=venue,
            organization=organization,
            created_by=request.user
        )

        # Add participants (Only players from the same organization)
        selected_participants = request.POST.getlist("participants")
        valid_participants = players.filter(id__in=selected_participants)  # Ensure only valid participants
        camp.participants.set(valid_participants)

        messages.success(request, "Camp/Tournament created successfully!")
        return redirect("camps_tournaments")  # Redirect after creation

    return render(request, "player_app/create_camp.html", {
        "organizations": organizations,  # Super Admin can select
        "players": players  # Filtered players for the user
    })


def delete_camp(request, camp_id):
    """
    Allows authorized users to soft-delete a camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)  # Remove organization filtering
    camp.is_deleted = True
    camp.save()
    # Log the deletion activity
    CampActivity.objects.create(
        camp=camp,
        action='deleted',
        performed_by=request.user,
        details=f"Camp/Tournament '{camp.name}' was deleted."
    )
    messages.success(request, 'Camp/Tournament deleted successfully.')
    return redirect('camps_tournaments')


def camp_detail(request, camp_id):
    """
    Displays details of a specific camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)
    return render(request, 'player_app/camp_detail.html', {'camp': camp})


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import CampTournament, Player, CampActivity


def edit_camp(request, camp_id):
    """
    Handles editing a specific camp/tournament, including participants.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)

    if request.method == 'POST':
        camp.name = request.POST.get('name')
        camp.camp_type = request.POST.get('camp_type')

        # Keep existing start date (only allow editing end date)
        end_date = request.POST.get('end_date')
        if end_date:
            camp.end_date = end_date

        camp.venue = request.POST.get('venue')

        # Update participants (only from the same organization)
        participant_ids = request.POST.getlist('participants')
        camp.participants.set(participant_ids)

        camp.save()

        # Log the update activity
        CampActivity.objects.create(
            camp=camp,
            action='updated',
            performed_by=request.user,
            details=f"Camp/Tournament '{camp.name}' was updated."
        )

        messages.success(request, "Camp/Tournament updated successfully!")
        return redirect('camp_detail', camp_id=camp.id)

    # Get only players from the same organization
    participants = Player.objects.filter(organization=camp.organization)

    return render(request, 'player_app/edit_camp.html', {
        'camp': camp,
        'participants': participants
    })


def trash_camps(request):
    """
    Lists all deleted camps/tournaments for management (restore or permanent delete).
    """
    deleted_camps = CampTournament.objects.filter(is_deleted=True)
    return render(request, 'player_app/trash_camps.html', {'deleted_camps': deleted_camps})


def restore_camp(request, camp_id):
    """
    Restores a soft-deleted camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id, is_deleted=True)
    camp.is_deleted = False
    camp.save()
    messages.success(request, 'Camp/Tournament restored successfully.')
    return redirect('trash_camps')


def permanently_delete_camp(request, camp_id):
    """
    Permanently deletes a camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id, is_deleted=True)
    camp.delete()
    messages.success(request, 'Camp/Tournament permanently deleted.')
    return redirect('trash_camps')


def add_participant(request, camp_id, participant_id):
    camp = get_object_or_404(CampTournament, id=camp_id)
    participant = get_object_or_404(Player, id=participant_id)

    camp.participants.add(participant)

    # Log the participant addition activity
    CampActivity.objects.create(
        camp=camp,
        action='player_added',
        performed_by=request.user,
        details=f"Player '{participant.name}' was added to the camp/tournament."
    )

    messages.success(request, f"Player '{participant.name}' added successfully.")
    return redirect('camp_detail', camp_id=camp.id)


def download_activity_history(request, camp_id):
    camp = get_object_or_404(CampTournament, id=camp_id)
    activities = camp.activities.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{camp.name}_activity_history.csv"'

    writer = csv.writer(response)
    writer.writerow(['Action', 'Performed By', 'Timestamp', 'Details'])

    for activity in activities:
        writer.writerow([
            activity.get_action_display(),
            activity.performed_by.username if activity.performed_by else "System",
            activity.timestamp,
            activity.details
        ])

    return response


def create_program(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        program_type = request.POST.get('program_type')
        template = 'template' in request.POST

        Program.objects.create(
            name=name,
            description=description,
            program_type=program_type,
            template=template,
            created_by=request.user
        )
        messages.success(request, "Program created successfully!")
        return redirect('program_list')

    return render(request, 'player_app/create_program.html')


def assign_program(request):
    """
    Assign a program to a player.
    """
    if request.method == 'POST':
        player_id = request.POST.get('player_id')
        program_id = request.POST.get('program_id')  # Use `program_id` instead of `id`
        injury_id = request.POST.get('injury_id', None)  # Optional for rehab programs

        # Validate the program_id
        if not program_id:
            messages.error(request, "Please select a valid program.")
            return redirect('assign_program')

        # Get the player and program objects
        player = get_object_or_404(Player, id=player_id)
        program = get_object_or_404(Program, program_id=program_id)  # Use `program_id`

        # Validate rehab programs require an injury ID
        if program.program_type == 'rehab' and not injury_id:
            messages.error(request, "Rehab programs must be assigned with an injury ID.")
            return redirect('assign_program')

        # Assign the program to the player
        AssignedProgram.objects.create(
            player=player,
            program=program,
            injury_id=injury_id,
            assigned_by=request.user
        )

        messages.success(request, "Program assigned successfully!")
        return redirect('program_list')

    # Fetch all players and programs for the dropdowns
    players = Player.objects.all()
    programs = Program.objects.all()
    return render(request, 'player_app/assign_program.html', {'players': players, 'programs': programs})


def save_workout_data(request, program_id):
    """
    Save workout data for a specific assigned program.
    Super Admins and Players can save data.
    """
    assigned_program = get_object_or_404(AssignedProgram, id=program_id)

    # Check if the user is a Player, Super Admin, or Staff with appropriate permission
    if not (
            request.user.is_superuser or
            hasattr(request.user, 'player') or
            (hasattr(request.user, 'staff') and request.user.staff.assign_program)
    ):
        return HttpResponse(
            "<h1 style='color: red; text-align: center;'>Permission Denied</h1>"
            "<p style='text-align: center;'>You do not have permission to save workout data.</p>",
            status=403
        )

    if request.method == 'POST':
        workout_details = request.POST.get('workout_details')

        # If the user is a Super Admin, use a placeholder or assign the workout to the first player
        player = getattr(request.user, 'player', None)
        if request.user.is_superuser and not player:
            player = assigned_program.player  # Assign to the player in the AssignedProgram

        # Save workout data
        WorkoutData.objects.create(
            assigned_program=assigned_program,
            player=player,
            workout_details=workout_details
        )

        messages.success(request, "Workout data saved successfully!")
        return redirect('program_list')  # Redirect to the program list or another appropriate page

    return render(request, 'player_app/save_workout_data.html', {'assigned_program': assigned_program})


def program_list(request):
    """
    Displays a list of training programs based on user roles.
    """
    user = request.user

    # Super Admin can filter programs by organization
    if user.is_superuser:
        organizations = Organization.objects.all()
        selected_org_id = request.GET.get('organization')
        if selected_org_id:
            programs = Program.objects.filter(created_by__organization=selected_org_id)
        else:
            programs = Program.objects.all()
    else:
        # Regular users see only their organization's programs
        programs = Program.objects.filter(created_by__organization=user.organization)
        organizations = None  # No need for dropdown

    return render(request, 'player_app/program_list.html', {
        'programs': programs,
        'organizations': organizations
    })


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import InjuryForm
from .models import Player, Organization

@login_required
def create_injury(request):
    """
    Create an injury record with organization-based player filtering.
    """
    organizations = None  # Initialize organizations variable

    # Super Admin: Can select an organization
    if request.user.is_superuser:
        organizations = Organization.objects.all()  # Fetch all organizations
        players = Player.objects.all()  # Show all players
    else:
        # Organization users can only select players from their own organization
        if hasattr(request.user, 'organization') and request.user.organization:
            players = Player.objects.filter(organization=request.user.organization)
        else:
            players = Player.objects.none()  # No players if no organization found

    if request.method == 'POST':
        form = InjuryForm(request.POST, request.FILES)

        if form.is_valid():
            injury = form.save(commit=False)

            # Auto-assign organization if the user is not a super admin
            if not request.user.is_superuser:
                injury.player.organization = request.user.organization

            injury.save()
            return redirect('injury_list')

    else:
        form = InjuryForm()

    return render(request, 'player_app/create_injury.html', {
        'form': form,
        'players': players,
        'organizations': organizations
    })



from django.shortcuts import render
from .models import Injury, Organization

def injury_list(request):
    injuries = Injury.objects.all()
    organizations = Organization.objects.all()  # Fetch all organizations
    selected_org_id = request.GET.get("organization", None)

    # Filter injuries by selected organization (if applicable)
    if selected_org_id:
        injuries = injuries.filter(player__organization_id=selected_org_id)

    return render(request, "player_app/injury_list.html", {
        "injuries": injuries,
        "organizations": organizations,
        "selected_org_id": int(selected_org_id) if selected_org_id else None
    })



from django.shortcuts import render, get_object_or_404
from .models import Player, Injury, AssignedProgram, WorkoutData, CampTournament


def player_injury_details(request, player_id):
    player = get_object_or_404(Player, id=player_id)
    injuries = Injury.objects.filter(player=player)
    assigned_programs = AssignedProgram.objects.filter(player=player)
    workout_data = WorkoutData.objects.filter(assigned_program__in=assigned_programs)

    # Fetch all camps/tournaments the player has participated in
    camps = CampTournament.objects.filter(participants=player)

    return render(request, 'player_app/player_injury_details.html', {
        'player': player,
        'injuries': injuries,
        'assigned_programs': assigned_programs,
        'workout_data': workout_data,
        'camps': camps,  # Send camp data to template
    })


def update_injury(request, pk):
    injury = get_object_or_404(Injury, pk=pk)
    if request.method == 'POST':
        form = InjuryForm(request.POST, request.FILES, instance=injury)
        if form.is_valid():
            if form.cleaned_data['status'] == 'closed' and not request.FILES.get('medical_documents'):
                return HttpResponse("Medical documentation is required to close the injury.")
            form.save()
            return redirect('injury_list')
    else:
        form = InjuryForm(instance=injury)
    return render(request, 'player_app/update_injury.html', {'form': form})


@login_required
def update_injury_status(request, pk):
    injury = get_object_or_404(Injury, pk=pk)

    # Permission check
    if not (request.user.is_superuser or
            (hasattr(request.user, 'staff') and request.user.staff.injury_tracking)):
        return HttpResponseForbidden("You don't have permission to update the injury status.")

    if request.method == 'POST':
        status = request.POST.get('status')

        # If closing, redirect to the document upload page
        if status == 'closed':
            return render(request, 'player_app/confirm_close.html', {'injury': injury})

        # Otherwise, just update the status
        injury.status = status
        injury.save()
        return redirect('injury_list')


@login_required
def confirm_close(request, pk):
    injury = get_object_or_404(Injury, pk=pk)

    # Permission check
    if not (request.user.is_superuser or
            (hasattr(request.user, 'staff') and request.user.staff.injury_tracking)):
        return HttpResponseForbidden("You don't have permission to update the injury status.")

    if request.method == 'POST':
        medical_documents = request.FILES.get('medical_documents')

        if medical_documents:
            injury.medical_documents = medical_documents
            injury.status = 'closed'
            injury.save()
            return redirect('injury_list')
        else:
            return HttpResponse("Medical documentation is required to close the injury.")
    return redirect('injury_list')


def add_treatment_recommendation(request, injury_id):
    # Check if the user is a Physio or Super Admin
    physio = Staff.objects.filter(user=request.user, role="physio").first()
    is_super_admin = request.user.is_superuser  # Check if the user is a Super Admin

    if not physio and not is_super_admin:
        messages.error(request, "You must be a Physio or Super Admin to recommend treatments.")
        return redirect("injury_list")

    injury = get_object_or_404(Injury, id=injury_id)
    player = injury.player  # Ensure the correct player is linked to the injury

    # Get existing treatment recommendations for the injury
    treatment_recommendations = TreatmentRecommendation.objects.filter(injury=injury)

    if request.method == "POST":
        form = TreatmentRecommendationForm(request.POST)
        if form.is_valid():
            recommendation = form.save(commit=False)
            recommendation.injury = injury  # Link treatment to the injury
            recommendation.player = player  # Ensure the correct player is assigned

            if is_super_admin:
                # Allow Super Admin to select Physio manually
                physio_id = request.POST.get('physio')  # Get selected physio ID from the form data
                if physio_id:
                    recommendation.physio = Staff.objects.get(id=physio_id)
                else:
                    messages.error(request, "Please select a Physio.")
                    return redirect("add_treatment", injury_id=injury.id)
            else:
                # Automatically assign the logged-in Physio
                recommendation.physio = physio

            recommendation.save()

            messages.success(request, "Treatment recommendation added successfully!")
            return redirect("player_injury_details", player_id=player.id)

    else:
        form = TreatmentRecommendationForm(initial={'player': player})  # Pre-fill the correct player

    return render(request, "player_app/add_treatment.html", {
        "form": form,
        "injury": injury,
        "treatment_recommendations": treatment_recommendations,
        "physios": Staff.objects.filter(role="physio"),  # Pass available Physios for selection
    })

# --------------------------------------------------------------------------------------------------------------------------------
# Team Management Views
# --------------------------------------------------------------------------------------------------------------------------------
from datetime import datetime, date, time
def teams_dashboard(request):
    category_keys = [
        ('boys_under-15', 'Boys under 15'),
        ('boys_under-19', 'Boys under 19'),
        ('men_under-23', 'Men Under 23'),
        ('men_senior', 'Men Senior'),
    ]

    categories = []
    for key, display in category_keys:
        teams = Team.objects.filter(category=key, is_active=True).prefetch_related('players', 'staff')
        categories.append((key, display, teams))

    today = date.today()

    end = CampTournament.objects.filter(
        organization=request.user.organization,
        is_deleted=False,
    )
   

    # Ongoing: started on or before today, and end_date after today or null (no end date means ongoing)
    ongoing_camps = CampTournament.objects.filter(
        organization=request.user.organization,
        is_deleted=False,
        start_date__lte=today,
    ).filter(
        Q(end_date__gt=today) | Q(end_date__isnull=True)
    ).order_by('name')

    # Closed: started on or before today and ended on or before today
    closed_camps = CampTournament.objects.filter(
        organization=request.user.organization,
        is_deleted=False,
        start_date__lte=today,
        end_date__lte=today,
    ).order_by('name')


    camps = {
        'ongoing': ongoing_camps,
        'closed': closed_camps,
    }

    context = {
        'categories': categories,
        'camps': camps,
    }
    return render(request, 'player_app/organization/teams_dashboard.html', context)
@login_required
def player_record(request):
    user_organization = getattr(request.user, 'organization', None)
    players_in_org = Player.objects.filter(organization=user_organization)
    session = None
    if 'report_settings' in request.session:
        session = True
    return render(request, 'player_app/organization/player_record.html', {
        'players': players_in_org,'session': session,
    })


@login_required
def player_data(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        player_id = request.POST.get('playerSelect')
        test = request.POST.get('testSelect')
        date_option = request.POST.get('date_option')
        num_tests = request.POST.get('numTestsToShow')  # Get the selected number of tests

        # Validate date range
        if date_option == 'range':
            if not start_date or not end_date:
                players_in_org = Player.objects.filter(organization=getattr(request.user, 'organization', None))
                error_message = "Please select both start and end dates for the date range."
                return render(request, 'player_app/organization/player_record.html', {
                    'players': players_in_org,
                    'error_message': error_message,
                })

        # Defaults for start and end dates
        if not start_date:
            start_date = '2000-01-01'
        if not end_date:
            from datetime import date
            end_date = date.today().strftime('%Y-%m-%d')

        # Pass num_tests parameter within redirect URL query string
        return redirect(f"{reverse('player_info', kwargs={'player_id': player_id, 'test': test, 'start': start_date, 'end': end_date})}?num_tests={num_tests}")
    else:
        players_in_org = Player.objects.filter(organization=getattr(request.user, 'organization', None))
        return render(request, 'player_app/organization/player_record.html', {
            'players': players_in_org,
        })
    
@login_required
def player_info(request, player_id, test, start, end):
    player = get_object_or_404(Player, id=player_id)
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)

    num_tests = int(request.GET.get('num_tests', 5))

    # Filter tests by player, test, date range
    if start and end:
        tests_qs = TestAndResult.objects.filter(player_id=player_id, test=test, date__range=[start, end])
    else:
        tests_qs = TestAndResult.objects.filter(player_id=player_id, test=test)

    tests = tests_qs.order_by('-date')[:num_tests]

    player_cat = player.age_category
    category = Category.objects.filter(name=player_cat).first() if player_cat else None

    target_value = None
    if category:
        target_obj = CategoryTarget.objects.filter(category=category).order_by('-settings__created_at').first()
        if target_obj:
            target_value = target_obj.target_value

    # Read min/max formula and toggles from session with defaults
    session_settings = request.session.get('report_settings', {})
    min_max_formula = session_settings.get('min_max_formula', 'all_players')
    min_is_better = session_settings.get('min_is_better', False)
    grp_avg_option = session_settings.get('grp_avg_option', None)
    
    # Calculate min and max based on formula
    if min_max_formula == "all_players":
        min_value = TestAndResult.objects.filter(test=test).aggregate(Min('best'))['best__min']
        max_value = TestAndResult.objects.filter(test=test).aggregate(Max('best'))['best__max']
       
    elif min_max_formula == "all_players_by_gender":
        min_value = TestAndResult.objects.filter(test=test, player__gender=player.gender).aggregate(Min('best'))['best__min']
        max_value = TestAndResult.objects.filter(test=test, player__gender=player.gender).aggregate(Max('best'))['best__max']
        
    elif min_max_formula == "category_based":
        min_value = TestAndResult.objects.filter(test=test, player__age_category=player.age_category).aggregate(Min('best'))['best__min']
        max_value = TestAndResult.objects.filter(test=test, player__age_category=player.age_category).aggregate(Max('best'))['best__max']
          
    elif min_max_formula == "date_based":
        group_filter = TestAndResult.objects.filter(test=test)
        if start and end:
            group_filter = group_filter.filter(date__range=[start, end])
        min_value = group_filter.aggregate(Min('best'))['best__min']
        max_value = group_filter.aggregate(Max('best'))['best__max']
    elif min_max_formula == "manual_entry":
        min_value = session_settings.get('manual_min', None)
        max_value = session_settings.get('manual_max', None)
    else:
        min_value = tests_qs.aggregate(Min('best'))['best__min']
        max_value = tests_qs.aggregate(Max('best'))['best__max']

    # Swap min and max if "min is better"
    if min_is_better and (min_value is not None and max_value is not None):
        min_value, max_value = max_value, min_value

    individual_avg = tests.aggregate(avg_best=Avg('best'))['avg_best']

    # Group average calculation based on grp_avg_option
    group_tests = TestAndResult.objects.filter(test=test)

    if grp_avg_option == "all_players_date":
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end])
    elif grp_avg_option == "all_players_gender_date":
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end], player__gender=player.gender)
        else:
            group_tests = group_tests.filter(player__gender=player.gender)
    elif grp_avg_option == "category_based_date":
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end], player__age_category=player.age_category)
        else:
            group_tests = group_tests.filter(player__age_category=player.age_category)
    
    elif grp_avg_option == "camp_or_tournament":
        if tests.exists():
            phase = tests.first().phase
            group_tests = TestAndResult.objects.filter(test=test, phase=phase)
            group_avg = group_tests.aggregate(avg_best=Avg('best'))['avg_best']
        else:
            group_avg = None

    else:
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end])

    group_avg = group_tests.aggregate(avg_best=Avg('best'))['avg_best']

    individual_normalized = (
        (individual_avg - min_value) / (max_value - min_value) * 100
        if max_value is not None and min_value is not None and max_value != min_value else 0
    )

    session = None
    if 'report_settings' in request.session:
        session = True

    abc = "start"

    return render(request, 'player_app/organization/player_record.html', {
        'player': player,
        'players': players,
        'tests': tests,
        'individual_avg': individual_avg,
        'group_avg': group_avg,
        'abc': abc,
        'max_value': max_value,
        'min_value': min_value,
        'test': test,
        'individual_normalized': individual_normalized,
        'target_value': target_value,
        'session': session,
    })
 


def delete_session(request):
    if 'report_settings' in request.session:
        del request.session['report_settings']
    return redirect('player_record')

@login_required
def get_players_by_test(request):
    test = request.GET.get('test')
    user_organization = getattr(request.user, 'organization', None)
    if not user_organization or not test:
        return JsonResponse({'players': []})

    # Get player IDs who have done the selected test and belong to the organization
    player_ids = TestAndResult.objects.filter(
        test=test, player__organization=user_organization
    ).values_list('player_id', flat=True).distinct()

    players = Player.objects.filter(id__in=player_ids).values('id', 'name')
    players_list = list(players)

    return JsonResponse({'players': players_list})

from django.utils.dateparse import parse_date
@login_required
def get_player_test_results(request):
    player_id = request.GET.get('player_id')
    test = request.GET.get('test')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    limit = int(request.GET.get('limit', 5))

    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None

    filters = {
        'player_id': player_id,
        'test': test
    }

    if start_date and end_date:
        filters['date__range'] = [start_date, end_date]

    qs = TestAndResult.objects.filter(**filters).order_by('-date')[:limit]

    results = list(qs.values('date', 'best', 'phase'))
    return JsonResponse({'results': results, 'count': qs.count()})

def new_test_details(request):
    results = TestAndResult.objects.select_related('player').all()
    return render(request, 'player_app/organization/new_test_details.html', {'results': results})



from django.forms.models import model_to_dict
def nomative_data(request):
    data = NomativeData.objects.all()
    return render(request, 'player_app/organization/nomative_data.html', {'datas': data})
# @login_required
# def new_test_dashboard(request):
#     user_organization = getattr(request.user, 'organization', None)
#     if not user_organization:
#         return render(request, 'player_app/organization/test_results.html', {
#             'error_message': "Your account is not linked to any organization.",
#         })
#     players_in_org = Player.objects.filter(organization=user_organization)

#     # Base queryset ordered by player, test and date
#     qs = TestAndResult.objects.select_related('player').filter(player__in=players_in_org).order_by(
#         'player__name', 'test', 'date'
#     )

#     # Calculate group averages per test using "best" field (one average per test)
#     group_averages = TestAndResult.objects.filter(player__in=players_in_org) \
#         .values('test').annotate(group_avg=Avg('best')).order_by('test')
#     group_avg_dict = {item['test']: item['group_avg'] for item in group_averages}

#     # Calculate individual averages per player per test using "best" field
#     indv_averages = TestAndResult.objects.filter(player__in=players_in_org).values('player', 'test').annotate(indv_avg=Avg('best'))
#     indv_avg_dict = {(item['player'], item['test']): item['indv_avg'] for item in indv_averages}

#     # Annotate each trial with averages for easy display in template
#     trials_with_avg = []
#     for trial in qs:
#         trial.individual_average = indv_avg_dict.get((trial.player.id, trial.test), None)
#         trial.group_average = group_avg_dict.get(trial.test, None)
#         trials_with_avg.append(trial)

#     context = {
#         'trials': trials_with_avg,
#     }

#     return render(request, 'player_app/organization/new_test_dashboard.html', context)

def report_settings_view(request):
    categories = Category.objects.all()

    # Get latest targets from DB for this user (latest ReportSettings)
    latest_settings = ReportSettings.objects.filter(user=request.user).order_by('-created_at').first()
    targets_dict = {}
    if latest_settings:
        targets_dict = {ct.category_id: ct.target_value for ct in latest_settings.category_targets.all()}

    # Get non-target settings from session
    session_settings = request.session.get('report_settings', {})

    # Build dummy settings for template with session data
    settings_data = type('Settings', (), {})()
    settings_data.min_max_formula = session_settings.get('min_max_formula', 'all_players')
    settings_data.min_is_better = session_settings.get('min_is_better', False)
    settings_data.indv_avg_option = session_settings.get('indv_avg_option', 'total_result')
    settings_data.grp_avg_option = session_settings.get('grp_avg_option', 'all_players_date')

    category_targets = []
    for category in categories:
        target_value = targets_dict.get(category.id, '')
        category_targets.append((category, target_value))

    context = {
        'category_targets': category_targets,
        'settings': settings_data,
    }
    return render(request, 'player_app/organization/report_settings.html', context)



def save_report_settings_view(request):
    if request.method != 'POST':
        return redirect(reverse('report_settings'))

    categories = Category.objects.all()

    # Capture form data from POST
    min_max_formula = request.POST.get('min_max_formula', 'all_players')
    min_is_better = 'min_is_better' in request.POST
    indv_avg_option = request.POST.get('indv_avg_option', 'total_result')
    grp_avg_option = request.POST.get('grp_avg_option', 'all_players_date')

    targets = {}
    for category in categories:
        val = request.POST.get(f'target_{category.id}', '').strip()
        if val:
            try:
                targets[str(category.id)] = float(val)
            except ValueError:
                # Skip invalid float entries or handle as needed
                pass

    # Save all to session
    request.session['report_settings'] = {
        'min_max_formula': min_max_formula,
        'min_is_better': min_is_better,
        'indv_avg_option': indv_avg_option,
        'grp_avg_option': grp_avg_option,
        'targets': targets,
    }

    # Mark session as modified to ensure Django saves changes
    request.session.modified = True

    messages.success(request, 'Report settings saved to session successfully!')
    return redirect(reverse('player_record'))

@login_required
def test_results_main(request):
    user = request.user
    # User's organization (adjust based on your user model)
    user_organization = getattr(user, 'organization', None)
    if not user_organization:
        return render(request, 'player_app/organization/test_dashboard.html', {
            'error_message': "Your account is not linked to any organization.",
        })

    # Players in user's organization
    players_in_org = Player.objects.filter(organization=user_organization)

    # Forms initialization with restricted player queryset
    if request.method == 'POST':
        add_form = TestAndResultForm(request.POST)
        add_form.fields['player'].queryset = players_in_org
        if add_form.is_valid():
            add_form.save()
            return redirect('test_dashboard')
    else:
        add_form = TestAndResultForm()
        add_form.fields['player'].queryset = players_in_org

    filter_form = TestSummaryFilterForm(request.GET or None)
    filter_form.fields['player'].queryset = players_in_org

    # Base queryset with related player, user; filtered by org players
    qs = (
        TestAndResult.objects
        .select_related('player', 'reported_by')
        .filter(player__in=players_in_org)
        .order_by('player__name', 'test', 'date', 'id')
    )

    # Apply filtering
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('player'):
            qs = qs.filter(player=filter_form.cleaned_data['player'])
        if filter_form.cleaned_data.get('test'):
            qs = qs.filter(test=filter_form.cleaned_data['test'])

    # Group test results by (player_id, test)
    grouped_results = defaultdict(list)
    for tr in qs:
        key = (tr.player.id, tr.test)
        grouped_results[key].append(tr)

    summary_rows = []
    for (player_id, test), trials in grouped_results.items():
        if not trials:
            continue

        # Sort trials by date and id to keep chronological order
        trials = sorted(trials, key=lambda x: (x.date or x.created_at, x.id))
        last_two_trials = trials[-2:] if len(trials) >= 2 else trials[-1:]

        # Trials data for columns
        trial_1_best = last_two_trials[0].best if len(last_two_trials) == 2 else None
        trial_2_best = last_two_trials[-1].best
        best_trial = min(t.best for t in trials if t.best is not None)

        last_trial_obj = last_two_trials[-1]

        # Individual average stored in PlayerAggregate or computed from TestAndResult
        player_agg = PlayerAggregate.objects.filter(player__id=player_id, test=test).first()
        indv_average = player_agg.individual_average if player_agg else None

        # Group average from GenderAggregate or CategoryAggregate fallback
        gender = getattr(last_trial_obj.player, 'gender', None)
        category = getattr(last_trial_obj.player, 'category', None)

        group_average = None
        if gender:
            gender_agg = GenderAggregate.objects.filter(gender=gender, test=test).first()
            group_average = gender_agg.average if gender_agg else None
        if group_average is None and category:
            cat_agg = CategoryAggregate.objects.filter(category=category, test=test).first()
            group_average = cat_agg.average if cat_agg else None

        summary_rows.append({
            'serial': None,  # Will assign later per test
            'player_name': last_trial_obj.player.name,
            'test': test,
            'last_date': last_trial_obj.date,
            'last_phase': last_trial_obj.phase,
            'trial_1': trial_1_best,
            'trial_2': trial_2_best,
            'best_trial': best_trial,
            'indv_average': indv_average,
            'group_average': group_average,
        })

    # Group rows by test for separate tables
    summary_by_test = defaultdict(list)
    for row in summary_rows:
        summary_by_test[row['test']].append(row)

    # Assign serial numbers per test group
    for test_name, rows in summary_by_test.items():
        for idx, row in enumerate(rows, start=1):
            row['serial'] = idx

    context = {
        'add_form': add_form,
        'form': filter_form,
        'summary_by_test': dict(summary_by_test),
    }

    return render(request, 'player_app/organization/test_result_main.html', context)


# New test Views
def test_dashboard_new(request):
    return render(request, 'player_app/organization/organization_main_test_dash.html')

from django.core.paginator import Paginator
from django.db.models import Q

def test_results_view(request, test_name):
    search_query = request.GET.get('search', '').strip()

    results = TestAndResult.objects.filter(test=test_name).select_related('player', 'reported_by').order_by('-date')

    # Filter by player name if search query is present
    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    from django.db.models import Max

    latest_dates = TestAndResult.objects.filter(
        player_id__in=player_ids,
        indv_average__isnull=False
    ).values('player_id').annotate(latest_date=Max('date'))

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = TestAndResult.objects.filter(
            player_id=entry['player_id'],
            date=entry['latest_date'],
            indv_average__isnull=False
        ).first()
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.indv_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,  # pass to template for form value retention
    }
    return render(request, 'player_app/organization/organization_test_data.html', context)


from django.shortcuts import render, redirect
from .models import TestAndResult, Player, User  # Make sure these models are imported

def add_test_results(request, test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    
     # Handle form submission
    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_id = request.POST.get('reported_by')

        # Basic field validation (add your own as needed)
        errors = []
        if not player_id: errors.append("Player is required.")
        if not test: errors.append("Test is required.")
        if not date: errors.append("Date is required.")
        if not phase: errors.append("Phase is required.")
        if not best: errors.append("Best is required.")
        if not reported_by_id: errors.append("Reported by is required.")


        # If no errors, save the result
        if not errors:
            phase_data = CampTournament.objects.get(id=int(phase))
            nomative_data = NomativeData.objects.get(final_level=float(best))
            total_distance = nomative_data.total_distance
            approximately_vo2max = nomative_data.approximately_vo2max

            player = Player.objects.get(pk=player_id)
            reported_by = User.objects.get(pk=reported_by_id)
            TestAndResult.objects.create(
                player=player,
                test=test,
                date=date,
                phase=phase_data,
                best=float(best),
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by
            )
            return redirect('test_results_by_name', test_name=test)
        # Pass errors back to template if any
        else:
            return render(request, 'player_app/organization/organization_test_add.html', {
                'test_name': test_name,
                'errors': errors,
                'players': players,
                'events': events,
                'staff':staff, 
                # you may need players/staff for dropdowns
            })
    else:
        return render(request, 'player_app/organization/organization_test_add.html', {
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Run A 3x6 test views
def add_run_3x6_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)


    return render(request, 'player_app/tests/runa3x6.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Glute Bridges Test views
def add_glute_bridges_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "S/L Glute Bridges"
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase = request.POST.get('phase')
        sl_right = request.POST.get('sl_glute_right')
        sl_left = request.POST.get('sl_glute_left')
        sl_difference = request.POST.get('sl_difference')
        slg_ratio = request.POST.get('slg_ratio')
        notes = request.POST.get('notes')
        reported_by_id = request.POST.get('reported_by')

        # Basic field validation (add your own as needed)
        errors = []
        if not player_id: errors.append("Player is required.")
        if not test: errors.append("Test is required.")
        if not date: errors.append("Date is required.")
        if not phase: errors.append("Phase is required.")
        if not sl_right: errors.append("S/L Glute Right is required.")
        if not sl_left: errors.append("S/L Glute Left is required.")
        if not sl_difference: errors.append("S/L Difference is required.")
        if not slg_ratio: errors.append("S/L Glute Ratio is required.")
        if not reported_by_id: errors.append("Reported by is required.")


        # If no errors, save the result
        if not errors:
            phase_data = CampTournament.objects.get(id=int(phase))
            

            player = Player.objects.get(pk=player_id)
            reported_by = User.objects.get(pk=reported_by_id)
            TestAndResult.objects.create(
                player=player,
                test=test,
                date=date,
                phase=phase_data,
                sl_right=sl_right,
                sl_left=sl_left,
                sl_difference=sl_difference,
                slg_ratio=slg_ratio,
                notes=notes,
                reported_by=reported_by
            )
            return redirect('test_results_by_name', test_name=test)
        # Pass errors back to template if any
        else:
            return render(request, 'player_app/organization/organization_test_add.html', {
                'test_name': test_name,
                'errors': errors,
                'players': players,
                'events': events,
                'staff':staff, 
                # you may need players/staff for dropdowns
            })
    return render(request, 'player_app/tests/glute_bridges.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Lunge Calf Raises Test views
def add_lunge_calf_raises_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)


    return render(request, 'player_app/tests/lunge_calf_raises.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# MB Rotational Throw Test views
def add_mb_rotational_throw_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)


    return render(request, 'player_app/tests/rotational_throws.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Copen Hagen Test views
def add_copen_hagen_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)


    return render(request, 'player_app/tests/copen_hagen.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# S/L Hop Test views
def add_sl_hop_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = "S/L Hop"

    return render(request, 'player_app/tests/sl_hop.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# CMJ Test views
def add_cmj_scores_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = "CMJ"

    return render(request, 'player_app/tests/cmj_scores.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

def add_anthropometry_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = "Anthropometry"

    return render(request, 'player_app/tests/anthropometry.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

def add_dexa_scan_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = "Dexa Scan"

    return render(request, 'player_app/tests/dexa_scan.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })


def add_blood_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)


    return render(request, 'player_app/tests/blood_work.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

def add_runa3_test(request,test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)


    return render(request, 'player_app/tests/runa3.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })